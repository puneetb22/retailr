import pandas as pd
import openpyxl

# Read the Excel file
df = pd.read_excel('attached_assets/shop_bill.xlsx')

# Get column info
print("Column names:", df.columns.tolist())
print("\nFirst 15 rows:")
print(df.head(15))

# Read sheet with openpyxl to preserve formatting
wb = openpyxl.load_workbook('attached_assets/shop_bill.xlsx')
sheet = wb.active
print("\nMerged cells:", sheet.merged_cells.ranges)

# Print cells with data to see placeholders
print("\nCells with placeholders:")
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        if cell.value and '{' in str(cell.value):
            print(f"Cell {cell.coordinate}: {cell.value}")
